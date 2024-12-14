from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.utils.timezone import now
from .models import Task, ChatMessage, User, Notification, Department, Designation,Company, Role, Permission
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from .utils import *
import openpyxl
from django.http import HttpResponse, HttpResponseForbidden
from functools import wraps
# from .consumers import send_notification

# Custom permission check decorator
def permission_required(permission_name):
    """Custom decorator to check if the user has a specific permission."""
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            # Check if the user has the required permission
            if not request.user.has_permission(permission_name):
                # Return a 403 Forbidden response if permission is not granted
                return render(request, '403.html')#HttpResponseForbidden("You don't have permission to access this page.")
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator

@permission_required('view_reports')
def some_protected_view(request):
    """A protected view where users need specific permissions to access it."""
    return render(request, 'protected_page.html')


# Create your views here.
def landingView(request):
    return render(request, 'landing_page.html')

# Registration View
def register_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        confirm_password = request.POST['confirm_password']

        if password != confirm_password:
            messages.error(request, "Passwords do not match!")
            return redirect('register')

        try:
            user = User.objects.create_user(username=username, email=email, password=password)
            messages.success(request, "Registration successful! Please login.")
            return redirect('login')
        except:
            messages.error(request, "Registration failed! Username or email may already exist.")
            return redirect('register')

    return render(request, 'register.html')


from django.contrib.auth import update_session_auth_hash
@login_required(login_url='/login/')
def user_profile(request):
    user = request.user

    if request.method == 'POST':
        # Handling the profile form
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        
        profile_picture = request.FILES.get('profile_picture')

        if first_name and last_name and email:
            # Update user info
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.phone_number = phone
            user.save()

            # Update profile picture if provided
            if profile_picture:
                user.profile_picture = profile_picture
                user.save()

            # Handle password change
            current_password = request.POST.get('current_password')
            new_password = request.POST.get('new_password')
            confirm_new_password = request.POST.get('confirm_new_password')

            if current_password and new_password and confirm_new_password:
                if user.check_password(current_password):
                    if new_password == confirm_new_password:
                        user.set_password(new_password)
                        user.save()
                        update_session_auth_hash(request, user)  # Keep user logged in after password change
                        messages.success(request, "Your password has been changed successfully.")
                    else:
                        messages.error(request, "New passwords do not match.")
                else:
                    messages.error(request, "Current password is incorrect.")

            messages.success(request, "Your profile has been updated successfully.")
            return redirect('user_profile')  # Avoid resubmission on refresh
        else:
            messages.error(request, "Please fill in all required fields.")
    else:
        profile_form = {
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'profile_picture': user.profile_picture
        }

    return render(request, 'userprofile.html', {
        'profile_form': profile_form,
        'user': user
    })


@login_required(login_url='/login/')
def create_employee(request):
    if request.method == "POST":
        username = request.POST.get('username')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        phone = request.POST.get('phone')
        email = request.POST.get('email')

        try:
            user = User.objects.create(username=phone, 
                                    first_name=first_name,
                                    last_name=last_name,
                                    email=email, 
                                    phone_number=phone,
                                    company=request.user.company)
            
            if 'profile_picture' in request.FILES:
                user.profile_picture = request.FILES['profile_picture']
                user.save()

            send_user_credentials(request, user, first_name, last_name)  # Email the credentials
            return redirect('employee_list')

        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return render(request, 'employee/create_employee.html')  
          
    return render(request, 'employee/create_employee.html')


@login_required(login_url='/login/')
def employee_list(request):
    employees = User.objects.filter(is_staff=False)  # Assuming employees are non-staff users
    return render(request, 'employee/employee_list.html', {'employees': employees})


# View Employee
@login_required(login_url='/login/')
def view_employee(request, id):
    employee = get_object_or_404(User, id=id, is_staff=False)  # Assuming employees are non-staff users
    return render(request, 'employee/view_employee.html', {'employee': employee})

# Edit Employee
@login_required(login_url='/login/')
def edit_employee(request, id):
    employee = get_object_or_404(User, id=id, is_staff=False)  # Assuming employees are non-staff users

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        phone_number = request.POST.get('phone_number')

        # Update fields
        employee.username = username
        employee.email = email
        employee.phone_number = phone_number

        if 'profile_picture' in request.FILES:
            employee.profile_picture = request.FILES['profile_picture']

        employee.save()
        messages.success(request, 'Employee details updated successfully.')
        return redirect('employee_list')

    return render(request, 'employee/edit_employee.html', {'employee': employee})

# Delete Employee
@login_required(login_url='/login/')
@permission_required('delete_employee')
def delete_employee(request, id):
    employee = get_object_or_404(User, id=id, is_staff=False)  # Assuming employees are non-staff users

    if request.method == 'POST':
        employee.delete()
        messages.success(request, 'Employee deleted successfully.')
        return redirect('employee_list')

    return render(request, 'employee/delete_employee.html', {'employee': employee})

@login_required(login_url='/login/')
def download_demo_file(request):
    # Create a sample Excel file (demo file)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    # Add headers to the Excel sheet
    ws.append(['First Name', 'Last Name', 'Email', 'Phone', 'Role'])# 'Department', 'Designation'])
    
    # Sample data
    employees = [
        ['Desh', 'Deepak', 'deshdeepakxpm@gmail.com', '7011101001', 'Founder'],# 'IT', 'CEO'],
        ['Himanshi', 'Kushwaha', 'radhikaxpm@gmail.com', '9717757483', 'Manager'],# 'HR', 'HR Specialist']
        ['Radhika', 'Kushwaha', 'radhikaxpm@gmail.com', '9717757483', 'Employee']# 'IT', 'Software Engineer']
    ]
    
    # Add sample data rows
    for employee in employees:
        ws.append(employee)
    
    # Prepare response with the demo file as an attachment
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employee_demo.xlsx'
    
    # Save the workbook to the HTTP response, which will automatically trigger a download in the browser
    wb.save(response)
    
    return response

@login_required(login_url='/login/')
# View for uploading the Excel file and creating bulk employees
def upload_employee_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            # Open the uploaded Excel file
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # Iterate over the rows and create employee records
            for row in ws.iter_rows(min_row=2, values_only=True):  # Start from the second row to skip headers
                first_name, last_name, email, phone, role = row #, department, designation = row
                # Create employee object (you can add validations here as needed)
                role_instance = Role.objects.create(name=role, company=request.user.company)
                employee = User(
                    username=phone,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    phone_number=phone,
                    role=role_instance,
                    company=request.user.company,
                    # department=department,
                    # designation=designation,
                )
                employee.save()
            
            messages.success(request, 'Employees uploaded successfully!')
        except Exception as e:
            messages.error(request, f"Error uploading file: {e}")
        
        return redirect('employee_list')

    return redirect('employee_list')


# Login View
def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            messages.success(request, "Login successful!")
            return redirect('dashboard')
        else:
            messages.error(request, "Invalid username or password!")
            return redirect('login')

    return render(request, 'login.html')


# Logout View
def logout_view(request):
    logout(request)
    messages.success(request, "Logged out successfully!")
    return redirect('login')


# Dashboard View (example protected view)
@login_required(login_url='/login/')
def dashboard_view(request):
    access_permissions = {
        'Settings': request.user.has_permission('Settings'),
        'Department': request.user.has_permission('Department'),
        'Designation': request.user.has_permission('Designation')
        }
    return render(request, 'dashboard.html', {'user': request.user,'access_permissions':access_permissions})



# Dashboard (Task List)
@login_required(login_url='/login/')
def task_list(request):
    tasks = Task.objects.filter().order_by('-created_at')
    tasks_message = ChatMessage.objects.filter().order_by('-id')
    return render(request, 'tasks/task_list.html', {'tasks': tasks, 'tasks_message':tasks_message})

# Task Creation
@login_required(login_url='/login/')
def create_task(request):
    if request.method == 'POST':
        title = request.POST['title']
        description = request.POST['description']
        eta = request.POST['eta']
        assigned_to_id = request.POST['assigned_to']
        assigned_to = User.objects.get(id=assigned_to_id)

        task = Task.objects.create(
            title=title,
            description=description,
            eta=eta,
            created_by=request.user,
            assigned_to=assigned_to
        )
        return redirect('task_list')

    users = User.objects.exclude(id=request.user.id,company=request.user.company,)
    return render(request, 'tasks/create_task.html', {'users': users})

# Task Detail and Chat
@login_required(login_url='/login/')
def task_detail(request, task_id):
    task = Task.objects.get(id=task_id)
    messages = ChatMessage.objects.filter(task=task).order_by('timestamp')
    return render(request, 'tasks/task_detail.html', {'task': task, 'messages': messages})


from django.views.decorators.csrf import csrf_exempt
import json

@csrf_exempt
def update_task_status(request, task_id):
    if request.method == 'POST':
        user = request.user
        task = get_object_or_404(Task, id=task_id)

        # Check if the user is the assigner or assignee
        # if user != task.user and user != task.assigned_to:
        if not request.user.has_perm('Change_Task_Status'):
            return JsonResponse({'success': False, 'message': 'Permission denied.'}, status=403)

        # Parse the new status from the request
        data = json.loads(request.body)
        new_status = data.get('status')

        # Permission checks
        if user == task.assigned_to:
            allowed_statuses = ['in_progress', 'completed', 'open', 'hold']
            if new_status not in allowed_statuses:
                return JsonResponse({'success': False, 'message': 'Invalid status update for assignee.'}, status=403)

        elif user == task.user:
            allowed_statuses = ['in_progress', 'completed', 'open', 'hold', 'reopen', 'closed']
            if new_status not in allowed_statuses:
                return JsonResponse({'success': False, 'message': 'Invalid status update for assigner.'}, status=403)

        # Update the task's status
        task.status = new_status
        task.save()

        # # Trigger notifications
        # send_notification(user, task, new_status)

        return JsonResponse({'success': True, 'message': 'Task status updated successfully.'})
    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=400)



def get_chat_messages(request, task_id):
    # Fetch messages for the given task from the database
    messages = ChatMessage.objects.filter(task_id=task_id).order_by('timestamp')

    # Prepare the data to send back as JSON
    message_data = []
    for message in messages:
        message_data.append({
            'sender': message.sender.username,  # Assuming sender is a User object
            'message': message.message,
            'timestamp': message.timestamp.strftime('%Y-%m-%d %H:%M:%S')
        })
    
    return JsonResponse({'messages': message_data})



# Task Chat View (to render chat and notifications page)
@login_required(login_url='/login/')
def task_chat_view(request, task_id):
    try:
        task = Task.objects.get(id=task_id)
    except Task.DoesNotExist:
        # Handle task not found
        return render(request, '404.html')  # Or redirect to some error page

    return render(request, 'chat_and_notifications.html', {'task': task})


# Save Chat Messages
from django.core.files.storage import default_storage
@login_required(login_url='/login/')
def save_message(request):
    if request.method == 'POST':
        task_id = request.POST['task_id']
        message = request.POST.get('message', '')
        file = request.FILES.get('file', None)

        task = Task.objects.get(id=task_id)
        chat_message = ChatMessage.objects.create(
            task=task,
            sender=request.user,
            message=message,
            file=file
        )

        response_data = {
            'status': 'Message saved',
            'sender': chat_message.sender.username,
            'message_id': chat_message.id,
            'timestamp': chat_message.timestamp,
            'file_url': chat_message.file.url if chat_message.file else None,
        }
        return JsonResponse(response_data)
    return JsonResponse({'status': 'Invalid request'}, status=400)



from django.views.decorators.csrf import csrf_exempt

@login_required(login_url='/login/')
@csrf_exempt
def delete_message_view(request, message_id):
    try:
        message = ChatMessage.objects.get(id=message_id, sender=request.user)
        message.is_deleted = True  # Soft delete
        message.message = "[Message deleted]"
        message.save()
        return JsonResponse({'status': 'success'})
    except ChatMessage.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Message not found'}, status=404)


@login_required(login_url='/login/')
def notifications_view(request):
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'notifications.html', {'notifications': notifications})


@login_required(login_url='/login/')
def mark_notification_as_read(request, notification_id):
    try:
        notification = Notification.objects.get(id=notification_id, user=request.user)
        notification.is_read = True
        notification.save()
        return JsonResponse({'status': 'success'})
    except Notification.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Notification not found'}, status=404)



from django.views.decorators.http import require_GET
@login_required(login_url='/login/')
@require_GET
def get_unread_message_count(request):
    task_id = request.GET.get('task_id')
    if task_id:
        count = ChatMessage.objects.filter(task_id=task_id, is_read=False).count()
        return JsonResponse({'unread_count': count})
    return JsonResponse({'error': 'Task ID is required'}, status=400)

@login_required(login_url='/login/')
@csrf_exempt
def mark_messages_as_read(request, task_id):
    if request.method == "POST":
        ChatMessage.objects.filter(task_id=task_id, is_read=False).update(is_read=True)
        return JsonResponse({'status': 'success'})
    return JsonResponse({'error': 'Invalid request method'}, status=400)










# List Departments
@login_required(login_url='/login/')
def list_departments(request):
    departments = Department.objects.filter(company=request.user.company)
    return render(request, 'department/list_departments.html', {'departments': departments})

# Create Department
@login_required(login_url='/login/')
def create_department(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        company = request.user.company

        Department.objects.create(name=name, company=company)
        messages.success(request, 'Department created successfully.')
        return redirect('list_departments')

    return render(request, 'department/create_department.html')

# Edit Department
@login_required(login_url='/login/')
def edit_department(request, id):
    department = get_object_or_404(Department, id=id, company=request.user.company)

    if request.method == 'POST':
        department.name = request.POST.get('name')
        department.save()
        messages.success(request, 'Department updated successfully.')
        return redirect('list_departments')

    return render(request, 'department/edit_department.html', {'department': department})

# Delete Department
@login_required(login_url='/login/')
@permission_required('delete_department')
def delete_department(request, id):
    department = get_object_or_404(Department, id=id, company=request.user.company)

    if request.method == 'POST':
        department.delete()
        messages.success(request, 'Department deleted successfully.')
        return redirect('list_departments')

    return render(request, 'department/delete_department.html', {'department': department})

# List Designations
@login_required(login_url='/login/')
def list_designations(request):
    designations = Designation.objects.filter(company=request.user.company)
    return render(request, 'designation/list_designations.html', {'designations': designations})

# Create Designation
@login_required(login_url='/login/')
def create_designation(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        company = request.user.company

        Designation.objects.create(name=name, company=company)
        messages.success(request, 'Designation created successfully.')
        return redirect('list_designations')

    return render(request, 'designation/create_designation.html')

# Edit Designation
@login_required(login_url='/login/')
def edit_designation(request, id):
    designation = get_object_or_404(Designation, id=id, company=request.user.company)

    if request.method == 'POST':
        designation.name = request.POST.get('name')
        designation.save()
        messages.success(request, 'Designation updated successfully.')
        return redirect('list_designations')

    return render(request, 'designation/edit_designation.html', {'designation': designation})

# Delete Designation
@login_required(login_url='/login/')
def delete_designation(request, id):
    designation = get_object_or_404(Designation, id=id, company=request.user.company)

    if request.method == 'POST':
        designation.delete()
        messages.success(request, 'Designation deleted successfully.')
        return redirect('list_designations')

    return render(request, 'designation/delete_designation.html', {'designation': designation})




# Role & Permissions

@login_required(login_url='/login/')
def role_list(request):
    """Display the list of roles."""
    roles = Role.objects.filter(company=request.user.company).order_by('-created_at')
    context = {
        'roles': roles,
    }
    return render(request, 'role_permissions/role_list.html', context)

@login_required(login_url='/login/')
def add_role(request):
    """Add a new role."""
    if request.method == 'POST':
        name = request.POST.get('name')
        permissions_ids = request.POST.getlist('permissions')
        company = request.user.company

        # Create the new role
        role = Role.objects.create(name=name, company=company, is_active=True)
        if permissions_ids:
            permissions = Permission.objects.filter(id__in=permissions_ids)
            role.permissions.set(permissions)
        messages.success(request, 'Role added successfully!')
        return redirect('role_list')

    # Fetch permissions for the form
    permissions = Permission.objects.filter(is_active=True)
    return render(request, 'role_permissions/add_role.html', {'permissions': permissions})

@login_required(login_url='/login/')
def edit_role(request, id):
    """Edit an existing role."""
    role = get_object_or_404(Role, id=id, company=request.user.company)

    if request.method == 'POST':
        role.name = request.POST.get('name')
        permissions_ids = request.POST.getlist('permissions')
        if permissions_ids:
            permissions = Permission.objects.filter(id__in=permissions_ids, is_active=True)
            role.permissions.set(permissions)
        else:
            role.permissions.clear()
        role.save()
        messages.success(request, 'Role updated successfully!')
        return redirect('role_list')

    # Fetch permissions for the form
    permissions = Permission.objects.filter(is_active=True)
    return render(request, 'role_permissions/edit_role.html', {'role': role, 'permissions': permissions})

@login_required(login_url='/login/')
@permission_required('delete_role')
def delete_role(request, id):
    """Delete a role."""
    role = get_object_or_404(Role, id=id, company=request.user.company)
    role.delete()
    messages.success(request, 'Role deleted successfully!')
    return redirect('role_list')



